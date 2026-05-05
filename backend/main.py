from fastapi import FastAPI, HTTPException, Form, Query
from database import banco_dados
import uuid
import re
import json
import io
import os
import hashlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import qrcode
from datetime import datetime, date

# uvicorn main:app --reload  

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], 
    allow_methods=["*"],
    allow_headers=["*"],
) 

qr_folder = "generated_qrcode"
os.makedirs(qr_folder, exist_ok=True)

def validar_email_backend(email: str):
    # Regex simples para validar formato de e-mail
    padrao = r"^[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*@(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?$"
    if not re.match(padrao, email.lower()):
        return False, "E-mail inválido."
    return True, ""

def validar_nome_sem_numeros(nome: str):
    # Verifica se existe algum dígito no nome
    if any(char.isdigit() for char in nome):
        return False, "O nome não pode conter números."
    if len(nome.strip()) < 3:
        return False, "Nome muito curto."
    return True, ""

@app.get("/")
async def read_index():
    return RedirectResponse(url="/docs")

@app.post("/usuarios/novo")
async def cadastro_user(nome: str, email: str, data_nasc: str, telefone: str = '', status_a: str = '', escola: str = "", curso_interesse: str = "", disciplina: str = ""):

    # Validação de Nome (sem números)
    v_nome, m_nome = validar_nome_sem_numeros(nome)
    if not v_nome: raise HTTPException(status_code=400, detail=m_nome)

    # Validação de E-mail
    v_email, m_email = validar_email_backend(email)
    if not v_email: raise HTTPException(status_code=400, detail=m_email)

    # Validação de Idade (15 anos)
    data_n_dt = datetime.strptime(data_nasc, "%Y-%m-%d")
    idade = (datetime.now() - data_n_dt).days // 365
    if idade < 15:
        raise HTTPException(status_code=400, detail="Você precisa ter pelo menos 15 anos.")

    u_user = str(uuid.uuid4())
    registro = datetime.now().isoformat()
    dados_user= {
        "id_user": u_user,
        "nome": nome.title(),
        "data_nasc": data_nasc,
        "email": email.lower().strip(),
        "telefone": telefone,
        "status_academico": status_a,
        "escola": escola,
        "curso_interesse": curso_interesse,
        "pontos": 0,
        "data_registro": registro,
        "disciplina": disciplina
    }

    try:
        # Tenta inserir no Supabase
        response = banco_dados.table("users").insert(dados_user).execute()
        return {"status": "Sucesso", "user": response.data[0]}

    except Exception as e:
        # Log do erro para você ver no terminal do VS Code
        print(f"Erro detectado: {e}")

        # Verifica se o erro é de chave duplicada (Código 23505 do Postgres)
        # O erro pode vir como string ou como objeto dependendo da lib
        error_msg = str(e)
        if "23505" in error_msg or "duplicate key" in error_msg:
            raise HTTPException(
                status_code=400, 
                detail="Este e-mail já está cadastrado em nossa base."
            )
        
        # Caso seja outro erro genérico
        raise HTTPException(
            status_code=500, 
            detail="Erro interno no servidor ao realizar cadastro."
        )

@app.get("/usuarios/verificar-admin")
async def verificar_admin(email: str):
    # Nota: use "users" (o nome que apareceu no seu INSERT)
    user = banco_dados.table('users').select("is_admin").eq("email", email).single().execute()
    
    if user.data:
        return {"is_admin": user.data.get('is_admin', False)}
    return {"is_admin": False}

@app.get("/usuarios/dados/exportar")
async def exportar_dados(
    data: date = Query(..., description="Data do relatório (formato: YYYY-MM-DD)"),
    formato: str = Query("json", description="Formato de exportação: 'json' ou 'xlsx'")
):
    # 1. Busca usuários registrados na data informada
    response = (
        banco_dados.table("users")
        .select("nome, pontos, disciplina")
        .gte("data_registro", f"{data}T00:00:00")
        .lte("data_registro", f"{data}T23:59:59")
        .execute()
    )

    if not response.data:
        raise HTTPException(
            status_code=404,
            detail=f"Nenhum usuário encontrado para a data {data}."
        )

    usuarios = response.data

    # 2. Exportar como JSON
    if formato == "json":
        relatorio = {
            "data_relatorio": str(data),
            "total_alunos": len(usuarios),
            "alunos": usuarios
        }

        json_bytes = json.dumps(relatorio, ensure_ascii=False, indent=2).encode("utf-8")

        return StreamingResponse(
            io.BytesIO(json_bytes),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=relatorio_{data}.json"}
        )

    # 3. Exportar como Excel (.xlsx)
    elif formato == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Relatório {data}"

        # Cabeçalho estilizado
        colunas = ["Nome", "Pontos", "Disciplina"]
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, titulo in enumerate(colunas, start=1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Preenche as linhas
        campos = ["nome", "pontos", "disciplina"]

        for row_idx, usuario in enumerate(usuarios, start=2):
            for col_idx, campo in enumerate(campos, start=1):
                ws.cell(row=row_idx, column=col_idx, value=usuario.get(campo, ""))

        # Ajusta largura das colunas automaticamente
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        # Salva em buffer e retorna
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=relatorio_{data}.xlsx"}
        )

    else:
        raise HTTPException(status_code=400, detail="Formato inválido. Use 'json' ou 'xlsx'.")
@app.get("/qrcodes/gerar")
async def gerar_qr(nome_local: str, pontos: int):
    dados_hash = f"{nome_local}-{pontos}{os.urandom(4).hex()}"
    code_hash = hashlib.sha256(dados_hash.encode()).hexdigest()[:12]

    # Inserção no banco com a pontuação correta
    banco_dados.table('qrcodes').insert({
        "code_hash": code_hash,
        "pontos": pontos,
        "local": nome_local
    }).execute()

    # 2. Gerar o QR Code
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(f"https://qr-code-hunt.vercel.app/scan/{code_hash}") # Link do seu front
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    # 3. Em vez de salvar em pasta, salva na memória (BytesIO)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0) # Volta para o início do "arquivo" na memória

    return StreamingResponse(buf, media_type="image/png")

@app.get("/qrcodes/download/{code_hash}")
async def download_qr(code_hash: str):
    # 1. Busca no Supabase se esse hash existe
    response = banco_dados.table("qrcodes").select("local").eq("code_hash", code_hash).execute()
    
    if not response.data:
        raise HTTPException(status_code=404, detail="QR Code não encontrado")

    # 2. Gera a imagem usando o hash que JÁ EXISTE
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    # IMPORTANTE: O link deve ser EXATAMENTE o que o seu front de scan espera
    qr.add_data(f"https://qr-code-hunt.vercel.app/scan/{code_hash}")
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)

    return StreamingResponse(buf, media_type="image/png")

@app.get("/qrcodes/listar")
async def listar_qrcodes():
    # Agora listamos apenas os que estão ativos
    response = banco_dados.table('qrcodes').select("*").execute()
    return response.data

@app.patch("/qrcodes/status/{code_hash}")
async def desativar_qr(code_hash: str):
    # Mudamos o status para false
    resultado = banco_dados.table('qrcodes').update({"ativo": False}).eq("code_hash", code_hash).execute()
    
    if not resultado.data:
        raise HTTPException(status_code=404, detail="QR Code não encontrado")
        
    return {"status": "sucesso", "mensagem": "QR Code desativado"}

@app.post("/capturar") 
async def capturar(user_id: str = Form(...), code_hash: str = Form(...)):
    try:
        # 1. Busca os dados do QR Code para saber quanto ele vale
        qr_data = banco_dados.table("qrcodes").select("pontos").eq("code_hash", code_hash).single().execute()
        
        if not qr_data.data:
            return {"status": "Erro", "msg": "QR Code não encontrado."}
        
        valor_pontos = qr_data.data['pontos']

        # 2. Registra a captura
        u_catch = str(uuid.uuid4())
        banco_dados.table("catch").insert({
            "id_catch": u_catch,
            "id_user": user_id,
            "catch_time": datetime.now().isoformat(),
            "code_hash": code_hash
        }).execute()

        # 3. RETORNA A PONTUAÇÃO REAL
        return {"status": "Sucesso", "pontos": valor_pontos}
    
    except Exception as e:
        # Tratamento de erro amigável que já configuramos
        msg = "Você já capturou este código!" if "duplicate" in str(e) else "Erro na captura."
        return {"status": "Erro", "msg": msg}
    
@app.get("/ranking")
async def ranking():
    # O segredo está no 'catch(count)': ele conta os registros relacionados na tabela catch
    res = banco_dados.table("users") \
        .select("id_user, nome, pontos, catch(count)") \
        .eq("is_admin", False) \
        .order("pontos", desc=True) \
        .limit(10) \
        .execute()
    
    # Formatando para o frontend receber uma lista limpa
    ranking_formatado = []
    for user in res.data:
        ranking_formatado.append({
            "id": user["id_user"],
            "nome": user["nome"],
            "pontos": user["pontos"],
            # Pega o count da lista retornada pelo Supabase
            "qrs_capturados": user["catch"][0]["count"] if user.get("catch") else 0
        })
        
    return ranking_formatado

@app.get("/login")
async def login(email: str, data_nasc: str):
    try:
        res = banco_dados.table("users")\
            .select("*")\
            .eq("email", email)\
            .eq("data_nasc", data_nasc)\
            .execute()
        
        if len(res.data) > 0:
            return {"user": res.data[0]}
        else:
            raise HTTPException(status_code=401, detail="E-mail ou data incorretos")
            
    except Exception as e:
        # Se o usuário digitar "123" e o banco esperar uma data, 
        # ele vai cair aqui em vez de derrubar o servidor.
        raise HTTPException(status_code=400, detail="Formato de data inválido. Use AAAA-MM-DD")